"""
Générateur de fichier Excel BC-compatible avec xmlMaps.xml complet.

Reproduit exactement la structure d'un export BC Configuration Package :
  - xl/xmlMaps.xml          : schéma XSD des tables/champs
  - xl/tables/tableN.xml    : liaisons colonnes ↔ XPath pour chaque onglet
  - xl/connections.xml      : connexion de données
  - Mise à jour rels + content_types

Utilise l'extension Talan QC pour les noms internes anglais (fieldInternalName,
tableEnglishName). Fallback automatique si ces champs sont absents.
"""
import io
import re
import zipfile
import openpyxl
from openpyxl.styles import PatternFill, Font as XLFont
from openpyxl.utils import get_column_letter

# ── Helpers de nommage ────────────────────────────────────────────────────────

def to_xml_name(s: str) -> str:
    """
    Convertit un nom BC (anglais) en nom d'élément XML valide.
    Règle BC : tiret → underscore, tous autres non-alphanum supprimés.
    Ex: "Gen. Prod. Posting Group" → "GenProdPostingGroup"
        "Rolled-up Material Cost"  → "Rolled_upMaterialCost"
    """
    s = str(s or "")
    s = s.replace("-", "_")
    s = re.sub(r"[^a-zA-Z0-9_]", "", s)
    if s and s[0].isdigit():
        s = "N" + s
    return s or "Field"


def to_unique_name(caption: str) -> str:
    """Caption français → uniqueName Excel (sans caractères spéciaux)."""
    return re.sub(r"[^a-zA-Z0-9]", "", str(caption or "")) or "Col"


_XSD_TYPES = {
    "Integer": "xsd:integer", "BigInteger": "xsd:integer",
    "Date": "xsd:date",       "Time": "xsd:time",
    "Boolean": "xsd:boolean", "DateTime": "xsd:dateTime",
}
_XML_DATA_TYPES = {
    "Integer": "integer", "BigInteger": "integer",
    "Date": "date",       "Time": "time",
}

def _xsd_type(dt: str)  -> str: return _XSD_TYPES.get(dt, "xsd:string")
def _xml_dtype(dt: str) -> str: return _XML_DATA_TYPES.get(dt, "string")


# ── Exemples et config style ──────────────────────────────────────────────────

_EXAMPLES = {
    18: {"N°": ["CLI-001","CLI-002"], "Nom": ["Société Test","Client Exemple"]},
    23: {"N°": ["FRN-001","FRN-002"], "Nom": ["Fournisseur Test","Autre"]},
    27: {"N°": ["ART-001","ART-002"], "Description": ["Article Test 1","Article Test 2"]},
}
_TYPE_EX = {
    "Code":"EX-001","Text":"Exemple","Decimal":"0.00",
    "Integer":"0","Date":"01/01/2025","Boolean":"Non",
}


# ── Étape 1 : Excel de base (openpyxl) ───────────────────────────────────────

def _build_base_xlsx(
    package: dict,
    tables_data: list[dict],
    options: dict,
) -> tuple[bytes, list[dict]]:
    """
    Génère le classeur avec openpyxl (sans xmlMaps — ajouté ensuite).
    Retourne (xlsx_bytes, sheets_meta) où sheets_meta = [{table, fields, sheet_idx}].
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    fill_blue  = PatternFill("solid", fgColor="FF2E6FBF")
    fill_dark  = PatternFill("solid", fgColor="FF1B3A6B")
    fill_meta  = PatternFill("solid", fgColor="FFEEEEEE")
    fill_desc  = PatternFill("solid", fgColor="FFFFF3CD")
    fill_ex    = PatternFill("solid", fgColor="FFF0FBF5")
    font_w     = XLFont(bold=True, color="FFFFFFFF", size=10)
    font_desc  = XLFont(italic=True, color="FF856404", size=9)
    font_ex    = XLFont(italic=True, color="FF0F6E56", size=9)
    font_meta  = XLFont(color="FF555555", size=9)

    try:
        from app.core.validator_axe_a import FIELD_DEFS
    except ImportError:
        FIELD_DEFS = {}

    sheets_meta = []
    pkg_code = package.get("code", "")

    for sheet_idx, table in enumerate(tables_data, 1):
        tid    = table.get("table_id", 0)
        tname  = table.get("table_name", str(tid))
        fields = [f for f in table.get("fields", [])
                  if options.get("include_custom_fields", True) or not f.get("is_custom")]
        if not fields:
            continue

        ws = wb.create_sheet(str(tname)[:31])
        fd_tbl  = FIELD_DEFS.get(str(tid), {})
        tbl_ex  = _EXAMPLES.get(tid, {})
        n_cols  = len(fields)

        # Row 1 : métadonnées BC
        ws.cell(1, 1, pkg_code); ws.cell(1, 2, tname); ws.cell(1, 3, tid)
        for c in range(1, n_cols + 1):
            ws.cell(1, c).fill = fill_meta
            ws.cell(1, c).font = font_meta

        # Row 3 : headers
        for i, f in enumerate(fields, 1):
            fname = f.get("field_name", "")
            fd    = fd_tbl.get(fname, {})
            req   = fd.get("req", False) or f.get("required", False)
            label = fname + (" *" if req and options.get("include_mandatory") else "")
            cell  = ws.cell(3, i, label)
            cell.fill = fill_dark if (req and options.get("include_mandatory")) else fill_blue
            cell.font = font_w

        next_row = 4

        # Row 4 (opt) : descriptions
        if options.get("include_descriptions"):
            for i, f in enumerate(fields, 1):
                fname = f.get("field_name", "")
                fd    = fd_tbl.get(fname, {})
                dtype = fd.get("type", "") or f.get("data_type", "")
                parts = [f"Type: {dtype}"] if dtype else []
                if fd.get("max"):   parts.append(f"Max: {fd['max']}")
                if fd.get("req") or f.get("required"): parts.append("OBLIGATOIRE")
                cell = ws.cell(next_row, i, " | ".join(parts))
                cell.fill = fill_desc; cell.font = font_desc
            next_row += 1

        # Rows exemples
        if options.get("include_examples"):
            for ex_idx in range(2):
                for i, f in enumerate(fields, 1):
                    fname = f.get("field_name", "")
                    fd    = fd_tbl.get(fname, {})
                    exs   = tbl_ex.get(fname, [])
                    val   = exs[ex_idx] if ex_idx < len(exs) else _TYPE_EX.get(fd.get("type",""), "")
                    cell  = ws.cell(next_row, i, val)
                    cell.fill = fill_ex; cell.font = font_ex
                next_row += 1

        # Largeurs colonnes
        for i, f in enumerate(fields, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(
                len(str(f.get("field_name", ""))), 12
            )

        sheets_meta.append({
            "sheet_idx": sheet_idx,
            "ws_title":  ws.title,
            "table_id":  tid,
            "table_name": tname,
            "table_english_name": table.get("table_english_name", tname),
            "fields": fields,
            "last_row": next_row,
        })

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), sheets_meta


# ── Étape 2 : Fichiers XML BC ─────────────────────────────────────────────────

def _build_xmlmaps(sheets_meta: list[dict]) -> bytes:
    """Génère xl/xmlMaps.xml à partir des métadonnées des onglets."""
    lines = [
        '<?xml version="1.0" encoding="utf-8"?>',
        '<x:MapInfo SelectionNamespaces="" xmlns:x="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
        '<Schema ID="Schema1">',
        '<xsd:schema xmlns:xsd="http://www.w3.org/2001/XMLSchema">',
        '<xsd:element name="DataList"><xsd:complexType><xsd:sequence>',
    ]
    for sm in sheets_meta:
        te_raw  = sm.get("table_english_name", sm["table_name"])
        te      = to_xml_name(te_raw)
        te_list = te + "List"
        lines += [
            f'<xsd:element name="{te_list}"><xsd:complexType><xsd:sequence>',
            '<xsd:element type="xsd:integer" name="TableID"></xsd:element>',
            '<xsd:element type="xsd:integer" name="PackageCode"></xsd:element>',
            f'<xsd:element name="{te}" maxOccurs="unbounded"><xsd:complexType><xsd:sequence>',
        ]
        for f in sm["fields"]:
            fin  = f.get("field_internal_name") or to_xml_name(f.get("field_name", ""))
            fname_xml = to_xml_name(fin)
            xsd  = _xsd_type(f.get("data_type", ""))
            lines.append(f'<xsd:element type="{xsd}" name="{fname_xml}"></xsd:element>')
        lines += [
            '</xsd:sequence></xsd:complexType></xsd:element>',
            '</xsd:sequence></xsd:complexType></xsd:element>',
        ]
    lines += [
        '</xsd:sequence></xsd:complexType></xsd:element>',
        '</xsd:schema></Schema>',
        '<Map ID="1" Name="DataList_Map" RootElement="DataList" SchemaID="Schema1" '
        'ShowImportExportValidationErrors="0" AutoFit="1" Append="0" '
        'PreserveSortAFLayout="1" PreserveFormat="1">',
        '<DataBinding FileBinding="1" ConnectionID="1" DataBindingLoadMode="1"></DataBinding>',
        '</Map></x:MapInfo>',
    ]
    return "".join(lines).encode("utf-8")


def _build_table_xml(sm: dict, table_num: int) -> bytes:
    """Génère xl/tables/tableN.xml pour un onglet."""
    fields   = sm["fields"]
    n_cols   = len(fields)
    last_col = get_column_letter(n_cols)
    last_row = sm.get("last_row", 4)
    ref      = f"A3:{last_col}{last_row}"
    te_raw   = sm.get("table_english_name", sm["table_name"])
    te       = to_xml_name(te_raw)
    te_list  = te + "List"

    lines = [
        '<?xml version="1.0" encoding="utf-8"?>',
        f'<x:table id="{table_num}" name="Table{table_num}" displayName="Table{table_num}" '
        f'ref="{ref}" tableType="xml" totalsRowShown="0" '
        'xmlns:x="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
        f'<x:autoFilter ref="{ref}"/>',
        f'<x:tableColumns count="{n_cols}">',
    ]
    seen_unique: dict[str, int] = {}
    for i, f in enumerate(fields, 1):
        caption  = f.get("field_name", f"Col{i}")
        fin      = f.get("field_internal_name") or to_xml_name(caption)
        fname_xml = to_xml_name(fin)
        xdt      = _xml_dtype(f.get("data_type", ""))
        xpath    = f"/DataList/{te_list}/{te}/{fname_xml}"
        uname    = to_unique_name(caption)
        # Dédupliquer le uniqueName
        if uname in seen_unique:
            seen_unique[uname] += 1
            uname = f"{uname}{seen_unique[uname]}"
        else:
            seen_unique[uname] = 0
        lines.append(
            f'<x:tableColumn id="{i}" uniqueName="{uname}" name="{caption}">'
            f'<x:xmlColumnPr mapId="1" xpath="{xpath}" xmlDataType="{xdt}"/>'
            f'</x:tableColumn>'
        )
    lines += [
        '</x:tableColumns>',
        '<x:tableStyleInfo name="TableStyleMedium2" showFirstColumn="0" '
        'showLastColumn="0" showRowStripes="1" showColumnStripes="0"/>',
        '</x:table>',
    ]
    return "".join(lines).encode("utf-8")


def _build_connections() -> bytes:
    """Génère xl/connections.xml."""
    return (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<x:connections xmlns:x="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<x:connection id="1" name="DataList_Map" type="4" refreshedVersion="0" background="1">'
        '<x:webPr xml="1" sourceData="1" url=""/>'
        '</x:connection>'
        '</x:connections>'
    ).encode("utf-8")


# ── Étape 3 : Fusion dans le zip ─────────────────────────────────────────────

NS_R   = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"
TYPE_TABLE    = f"{NS_R}/table"
TYPE_XMLMAPS  = f"{NS_R}/xmlMaps"
TYPE_CONNECTIONS = f"{NS_R}/connections"
CT_TABLE      = "application/vnd.openxmlformats-officedocument.spreadsheetml.table+xml"
CT_XMLMAPS    = "application/xml"
CT_CONNECTIONS = "application/vnd.openxmlformats-officedocument.spreadsheetml.connections+xml"


def _add_table_parts(sheet_xml: bytes, rel_id: str) -> bytes:
    """Injecte <tableParts> dans le worksheet XML."""
    s = sheet_xml.decode("utf-8")
    # Ajouter xmlns:r si absent
    if "xmlns:r=" not in s:
        s = s.replace(
            'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"',
            'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
            ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"',
            1
        )
    tp = f'<tableParts count="1"><tablePart r:id="{rel_id}"/></tableParts>'
    s  = s.replace("</worksheet>", tp + "</worksheet>", 1)
    return s.encode("utf-8")


def _sheet_rels_xml(table_num: int, rel_id: str) -> bytes:
    """Génère xl/worksheets/_rels/sheetN.xml.rels."""
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="{NS_PKG}">'
        f'<Relationship Id="{rel_id}" Type="{TYPE_TABLE}" '
        f'Target="../tables/table{table_num}.xml"/>'
        '</Relationships>'
    ).encode("utf-8")


def _merge_into_xlsx(
    base_bytes: bytes,
    sheets_meta: list[dict],
) -> bytes:
    """Assemble le xlsx final avec tous les fichiers XML BC."""
    xmlmaps_bytes = _build_xmlmaps(sheets_meta)
    connections_bytes = _build_connections()

    in_buf  = io.BytesIO(base_bytes)
    out_buf = io.BytesIO()

    # Map: sheet_idx → filename dans le zip (ex: xl/worksheets/sheet1.xml)
    # openpyxl génère sheet1.xml, sheet2.xml... dans l'ordre de création
    # On suppose que les onglets sont dans l'ordre de sheets_meta

    with zipfile.ZipFile(in_buf, "r") as zin:
        existing_names = set(zin.namelist())

        # Identifier les fichiers worksheet dans l'ordre
        sheet_files = sorted(
            [n for n in existing_names if re.match(r"xl/worksheets/sheet\d+\.xml$", n)],
            key=lambda x: int(re.search(r"\d+", x).group())
        )

        with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
            table_content_types = []
            table_rels_to_add   = []   # (rels_path, content)
            modified_sheets     = {}   # {filename: modified_bytes}

            # Modifier chaque worksheet pour ajouter <tableParts>
            for meta_idx, sm in enumerate(sheets_meta):
                if meta_idx >= len(sheet_files):
                    break
                sheet_file = sheet_files[meta_idx]
                table_num  = meta_idx + 1
                rel_id     = f"rIdTable{table_num}"

                # Modifier le worksheet XML
                original = zin.read(sheet_file)
                modified_sheets[sheet_file] = _add_table_parts(original, rel_id)

                # Table XML
                table_xml   = _build_table_xml(sm, table_num)
                table_fname = f"xl/tables/table{table_num}.xml"
                zout.writestr(table_fname, table_xml)
                table_content_types.append(
                    f'<Override PartName="/{table_fname}" ContentType="{CT_TABLE}"/>'
                )

                # Rels du worksheet
                sheet_num  = table_num  # sheet1 → table1
                rels_path  = f"xl/worksheets/_rels/sheet{sheet_num}.xml.rels"
                rels_bytes = _sheet_rels_xml(table_num, rel_id)
                # Si des rels existent déjà, on les fusionne
                if rels_path in existing_names:
                    existing_rels = zin.read(rels_path).decode("utf-8")
                    new_rel = (
                        f'<Relationship Id="{rel_id}" Type="{TYPE_TABLE}" '
                        f'Target="../tables/table{table_num}.xml"/>'
                    )
                    if rel_id not in existing_rels:
                        rels_bytes = existing_rels.replace(
                            "</Relationships>", new_rel + "</Relationships>"
                        ).encode("utf-8")
                table_rels_to_add.append((rels_path, rels_bytes))

            # Écrire les fichiers originaux (avec les modifs)
            for item in zin.infolist():
                fname = item.filename
                if fname in modified_sheets:
                    zout.writestr(item, modified_sheets[fname])
                elif fname in {p for p, _ in table_rels_to_add}:
                    pass  # sera écrit plus bas
                elif fname == "xl/_rels/workbook.xml.rels":
                    s = zin.read(fname).decode("utf-8")
                    additions = ""
                    if "xmlMaps" not in s:
                        additions += (
                            f'<Relationship Id="rIdXmlMaps" Type="{TYPE_XMLMAPS}" '
                            'Target="xmlMaps.xml"/>'
                        )
                    if "connections" not in s:
                        additions += (
                            f'<Relationship Id="rIdConnections" Type="{TYPE_CONNECTIONS}" '
                            'Target="connections.xml"/>'
                        )
                    if additions:
                        s = s.replace("</Relationships>", additions + "</Relationships>")
                    zout.writestr(item, s.encode("utf-8"))
                elif fname == "[Content_Types].xml":
                    s = zin.read(fname).decode("utf-8")
                    ct_additions = ""
                    if "xmlMaps.xml" not in s:
                        ct_additions += f'<Override PartName="/xl/xmlMaps.xml" ContentType="{CT_XMLMAPS}"/>'
                    if "connections.xml" not in s:
                        ct_additions += f'<Override PartName="/xl/connections.xml" ContentType="{CT_CONNECTIONS}"/>'
                    for ct in table_content_types:
                        part_name = re.search(r'PartName="([^"]+)"', ct)
                        if part_name and part_name.group(1) not in s:
                            ct_additions += ct
                    if ct_additions:
                        s = s.replace("</Types>", ct_additions + "</Types>")
                    zout.writestr(item, s.encode("utf-8"))
                else:
                    zout.writestr(item, zin.read(fname))

            # Ajouter les rels des worksheets
            for rels_path, rels_bytes in table_rels_to_add:
                zout.writestr(rels_path, rels_bytes)

            # Ajouter xmlMaps et connections
            zout.writestr("xl/xmlMaps.xml", xmlmaps_bytes)
            zout.writestr("xl/connections.xml", connections_bytes)

    return out_buf.getvalue()


# ── Point d'entrée principal ──────────────────────────────────────────────────

def generate_bc_excel(
    package: dict,
    tables_data: list[dict],
    options: dict,
) -> bytes:
    """
    Génère un fichier Excel BC-compatible avec xmlMaps.xml complet.

    Args:
        package     : {"code": "003K-ARTICLE", "packageName": "..."}
        tables_data : retour de build_tables_data_for_export() — doit inclure
                      table_english_name et field_internal_name pour un xmlMaps exact.
                      Fallback automatique vers to_xml_name(field_name) si absent.
        options     : {
            include_mandatory, include_descriptions,
            include_examples, include_custom_fields
        }

    Returns:
        bytes du fichier .xlsx prêt pour st.download_button
    """
    base_bytes, sheets_meta = _build_base_xlsx(package, tables_data, options)
    return _merge_into_xlsx(base_bytes, sheets_meta)
