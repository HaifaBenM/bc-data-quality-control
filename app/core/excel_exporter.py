"""
Génération de templates Excel compatibles BC Configuration Package.

Structure de sortie par onglet (fidèle au format BC) :
    Ligne 1 : PackageCode | NomTable | TableID
    Ligne 2 : vide
    Ligne 3 : Headers (champs)
    Ligne 4 : [EXEMPLE - optionnel, à supprimer]
    Ligne 5+ : vide — à remplir par le client

Onglet supplémentaire « GUIDE » :
    Table | Champ | Libellé | Type | Obligatoire | Champ perso | Exemple
"""
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# ── Couleurs ──────────────────────────────────────────────────────────────────
_C = {
    "header_std":      "FF2E6FBF",   # bleu standard
    "header_required": "FF1B3A6B",   # bleu foncé pour obligatoires
    "header_custom":   "FFFFA500",   # orange pour champs perso (≥50000)
    "metadata_bg":     "FFEEEEEE",   # gris ligne 1
    "example_bg":      "FFF5F5F5",   # gris clair ligne exemple
    "example_txt":     "FF999999",   # texte gris exemple
    "guide_header":    "FF1B3A6B",   # guide header
    "required_mark":   "FFCC0000",   # marqueur obligatoire dans guide
    "white":           "FFFFFFFF",
}

# ── Exemples par type BC (fallback si non fourni) ─────────────────────────────
_TYPE_EXAMPLES = {
    "Code":     "EX-001",
    "Text":     "Exemple texte",
    "Decimal":  "100.00",
    "Integer":  "1",
    "Date":     "01/01/2025",
    "Boolean":  "Oui",
    "Option":   "Option1",
    "GUID":     "{00000000-0000-0000-0000-000000000000}",
}

# ── Exemples spécifiques par table/champ (enrichir progressivement) ───────────
_FIELD_EXAMPLES: dict[int, dict[str, str]] = {
    18: {  # Customer
        "No.":                    "CLI-001",
        "Name":                   "Société Exemple SAS",
        "Customer Posting Group": "NATIONAL",
        "Gen. Bus. Posting Group": "NATIONAL",
        "Payment Terms Code":     "30 JOURS",
        "Country/Region Code":    "FR",
        "VAT Registration No.":   "FR12345678901",
    },
    23: {  # Vendor
        "No.":                    "FRN-001",
        "Name":                   "Fournisseur Exemple SAS",
        "Vendor Posting Group":   "NATIONAL",
        "Gen. Bus. Posting Group": "NATIONAL",
        "Payment Terms Code":     "30 JOURS",
    },
    27: {  # Item
        "No.":               "ART-001",
        "Description":       "Article exemple",
        "Item Category Code": "MATIERES",
        "Base Unit of Measure": "PCS",
        "Unit Price":        "10.00",
    },
}


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _font(bold=False, color="FF000000", italic=False, size=11) -> Font:
    return Font(bold=bold, color=color, italic=italic, size=size)


def _thin_border() -> Border:
    s = Side(style="thin", color="FFD0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def _build_guide_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    tables_data: list[dict],
) -> None:
    headers = ["Table", "Champ (technique)", "Libellé", "Type BC",
               "Obligatoire", "Champ perso", "Exemple"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = _fill(_C["guide_header"])
        cell.font = _font(bold=True, color=_C["white"])
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22
    row = 2
    for table in tables_data:
        for field in table["fields"]:
            ws.cell(row, 1, table["table_name"])
            ws.cell(row, 2, field.get("field_name", ""))
            ws.cell(row, 3, field.get("field_caption") or field.get("field_name", ""))
            ws.cell(row, 4, field.get("data_type", ""))
            req_cell = ws.cell(row, 5, "Oui" if field.get("required") else "")
            if field.get("required"):
                req_cell.font = _font(bold=True, color=_C["required_mark"])
            ws.cell(row, 6, "Oui" if field.get("is_custom") else "")
            ws.cell(row, 7, field.get("example") or "")
            row += 1

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = [
            22, 28, 28, 14, 12, 12, 24
        ][col - 1]


def generate_package_template(
    package: dict,
    tables_data: list[dict],
    options: dict,
) -> bytes:
    """
    Génère un fichier Excel template pour le package BC.

    Args:
        package    : {code, packageName}
        tables_data: [
            {
              table_id  : int,
              table_name: str,
              fields    : [
                {field_no, field_name, field_caption, data_type,
                 required, is_custom, validate_field, example, description}
              ]
            }
        ]
        options    : {
            include_descriptions : bool,
            include_examples     : bool,
            include_custom_fields: bool,
            role                 : "consultant" | "client",
        }

    Returns:
        bytes — fichier .xlsx prêt pour st.download_button
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    guide_ws = wb.create_sheet("GUIDE")

    processed_tables = []

    for table in tables_data:
        fields = list(table["fields"])

        # Filtre champs perso
        if not options.get("include_custom_fields", True):
            fields = [f for f in fields if not f.get("is_custom", False)]

        # Filtre rôle client : champs obligatoires uniquement
        if options.get("role") == "client":
            fields = [f for f in fields if f.get("required") or f.get("validate_field")]

        if not fields:
            continue

        table_id = table.get("table_id", 0)
        sheet_name = str(table.get("table_name", f"Table{table_id}"))[:31]

        # Éviter les noms d'onglets dupliqués
        existing = [ws.title for ws in wb.worksheets]
        if sheet_name in existing:
            sheet_name = sheet_name[:28] + f"_{table_id}"

        ws = wb.create_sheet(sheet_name)
        pkg_code = package.get("code", "")

        # ── Ligne 1 : métadonnées ─────────────────────────────────────────────
        for col in range(1, len(fields) + 1):
            ws.cell(1, col).fill = _fill(_C["metadata_bg"])
            ws.cell(1, col).font = _font(size=9, color="FF555555")

        ws.cell(1, 1, pkg_code)
        ws.cell(1, 2, table.get("table_name", ""))
        ws.cell(1, 3, table_id)
        ws.row_dimensions[1].height = 16

        # ── Ligne 2 : vide ────────────────────────────────────────────────────

        # ── Ligne 3 : headers ─────────────────────────────────────────────────
        ws.row_dimensions[3].height = 20
        for col_idx, field in enumerate(fields, 1):
            caption = field.get("field_caption") or field.get("field_name", "")
            is_req  = field.get("required", False)
            is_cust = field.get("is_custom", False)

            label = f"{caption} *" if is_req else caption
            cell  = ws.cell(3, col_idx, label)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=False
            )

            if is_cust:
                cell.fill = _fill(_C["header_custom"])
                cell.font = _font(bold=True, color="FF000000")
            elif is_req:
                cell.fill = _fill(_C["header_required"])
                cell.font = _font(bold=True, color=_C["white"])
            else:
                cell.fill = _fill(_C["header_std"])
                cell.font = _font(color=_C["white"])

            # Commentaire description
            if options.get("include_descriptions"):
                desc = field.get("description", "")
                if not desc:
                    desc = f"Type : {field.get('data_type','')}"
                    if is_req:
                        desc += " · OBLIGATOIRE"
                    if is_cust:
                        desc += " · Champ personnalisé"
                cell.comment = Comment(desc, "BC QC Tool", height=80, width=200)

            # Largeur colonne
            ws.column_dimensions[get_column_letter(col_idx)].width = max(
                len(label) + 4, 14
            )

        # ── Ligne 4 : exemples ────────────────────────────────────────────────
        if options.get("include_examples"):
            ex_row = 4
            ws.cell(ex_row, 1, "[EXEMPLE — supprimer cette ligne]")
            ws.cell(ex_row, 1).font = _font(italic=True, color=_C["example_txt"], size=9)

            tbl_examples = _FIELD_EXAMPLES.get(table_id, {})
            for col_idx, field in enumerate(fields, 1):
                example = (
                    field.get("example")
                    or tbl_examples.get(field.get("field_name", ""))
                    or _TYPE_EXAMPLES.get(field.get("data_type", ""), "")
                )
                cell = ws.cell(ex_row, col_idx, example)
                cell.fill  = _fill(_C["example_bg"])
                cell.font  = _font(italic=True, color=_C["example_txt"], size=9)
                cell.alignment = Alignment(horizontal="left")

        # Freeze pane sur ligne 3 (headers)
        ws.freeze_panes = "A4"

        processed_tables.append({**table, "fields": fields})

    # ── Guide ─────────────────────────────────────────────────────────────────
    _build_guide_sheet(guide_ws, processed_tables)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
