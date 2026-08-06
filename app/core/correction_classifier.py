"""
Classification des anomalies "Code de référence invalide" (et "Souches de
n° non résolvable") en deux catégories :

- VALEUR_CORRIGIBLE   : le code saisi est proche d'un code qui existe déjà
                        dans la table référencée BC (faute de frappe probable,
                        ou mauvais code choisi) -> corrigible dans le fichier.
- PREALABLE_BC_REQUIS : aucun code valide ne ressemble à la valeur saisie ->
                        le code n'existe tout simplement pas côté BC. Aucune
                        valeur saisie dans le fichier ne sera valide tant que
                        cette donnée maîtresse n'est pas créée dans BC.

Distinction basée sur une comparaison floue (difflib, stdlib — aucune
dépendance supplémentaire) contre l'ensemble des codes valides déjà récupéré
par Axe B (get_reference_values_by_table_id) : aucun appel BC additionnel.
"""
from __future__ import annotations
import difflib
import io
import pandas as pd

# Score de similarité minimal pour considérer un code existant comme "faute
# de frappe probable" plutôt que "code inexistant". Pas encore calibré sur
# un jeu de cas réel BC — à ajuster si trop de faux positifs/négatifs
# apparaissent en pratique (valeur de départ raisonnable, pas une vérité
# mesurée).
FUZZY_MATCH_THRESHOLD = 0.72
MAX_SUGGESTIONS = 3


def suggest_close_codes(
    value: str, valid_codes: set[str], limit: int = MAX_SUGGESTIONS
) -> list[tuple[str, float]]:
    """Codes valides les plus proches de `value`, triés par score décroissant."""
    if not value or not valid_codes:
        return []
    scored = []
    for code in valid_codes:
        if not code:
            continue
        ratio = difflib.SequenceMatcher(None, value.lower(), str(code).lower()).ratio()
        if ratio >= FUZZY_MATCH_THRESHOLD:
            scored.append((code, ratio))
    scored.sort(key=lambda x: x[1], reverse=True)
    return scored[:limit]


def classify_reference_anomaly(value: str, valid_codes: set[str]) -> dict:
    """
    Classifie une anomalie "Code de référence invalide".

    Retourne :
      {
        "classification": "VALEUR_CORRIGIBLE" | "PREALABLE_BC_REQUIS",
        "suggestions": [(code, score), ...],  # non vide seulement si VALEUR_CORRIGIBLE
      }
    """
    suggestions = suggest_close_codes(value, valid_codes)
    if suggestions:
        return {"classification": "VALEUR_CORRIGIBLE", "suggestions": suggestions}
    return {"classification": "PREALABLE_BC_REQUIS", "suggestions": []}


def build_prerequisites_report(
    anomalies: list[dict], profile_code: str = "", company_id: str = "",
) -> list[dict]:
    """
    Extrait les anomalies PREALABLE_BC_REQUIS et les regroupe par
    (table référencée, valeur manquante) pour produire une checklist de
    données maîtresses à créer côté BC avant import — distincte du fichier
    corrigé (ce ne sont PAS des corrections de valeur).

    Nom de table : interroge BC dynamiquement en priorité (cache Supabase +
    endpoint AL Table Caption API) si profile_code/company_id sont fournis ;
    ne retombe sur le dictionnaire statique master_data_config que si
    l'appel BC échoue ou si les identifiants ne sont pas fournis (usage sans
    contexte BC, ex: tests).
    """
    from app.core.master_data_config import get_table_label

    grouped: dict[tuple, dict] = {}
    for a in anomalies:
        if a.get("Classification") != "PREALABLE_BC_REQUIS":
            continue
        table_id = str(a.get("Table référencée", ""))
        key = (table_id, a.get("Valeur", ""))
        if key not in grouped:
            table_name = None
            if profile_code and company_id and table_id:
                try:
                    from app.db.metadata_db import get_table_caption_cached
                    table_name = get_table_caption_cached(profile_code, company_id, table_id)
                except Exception:
                    table_name = None  # BC injoignable — on retombe sur le statique juste en dessous
            if not table_name:
                table_name = get_table_label(table_id)

            grouped[key] = {
                "Table référencée BC": table_id,
                "Nom table BC":        table_name,
                "Code manquant":       a.get("Valeur", ""),
                "Champs concernés":    set(),
                "Occurrences":         0,
            }
        grouped[key]["Champs concernés"].add(a.get("Champ", ""))
        grouped[key]["Occurrences"] += 1

    report = []
    for row in grouped.values():
        row["Champs concernés"] = ", ".join(sorted(c for c in row["Champs concernés"] if c))
        report.append(row)
    return sorted(report, key=lambda r: -r["Occurrences"])


_PREREQ_COLUMNS = [
    "Table référencée BC", "Nom table BC", "Code manquant",
    "Champs concernés", "Occurrences",
]


# ── Contrôle croisé GL Account <-> groupes comptables (socle MDD Compta) ────
#
# Confirmé par test réel le 23/07/2026 : importer 92/93/94 après GL Account
# échoue quand même si le compte général référencé par un champ-compte de
# 92/93/94 n'a pas LUI-MÊME ses propres Groupe compta. marché / Groupe
# compta. produit renseignés — erreur BC "Groupe compta. produit doit avoir
# une valeur dans Compte général: N°=<compte>. Il ne peut pas être vide ou
# nul." Ce n'est pas une histoire de table manquante (le compte existe déjà,
# GL Account est intégré) : c'est un champ précis, sur une ligne précise,
# manquant.
#
# Vérification 100% inter-onglets DU MÊME fichier déposé, aucun appel BC :
# le compte et le groupe qui le référence sont dans le même package MDD
# Comptabilité (socle figé confirmé par Rami/Bilel), donc les deux onglets
# sont déjà là au moment du contrôle.

GL_ACCOUNT_REQUIRED_FIELDS = ["Groupe compta. marché", "Groupe compta. produit"]

_ACCOUNT_FIELD_PREFIXES = ("compte", "cpte")
_ACCOUNT_FIELD_EXCLUSIONS = {
    "afficher tous les comptes lors de la consultation",
}

# REMPLACÉ PAR UNE RÈGLE DYNAMIQUE (04/08/2026) — voir
# _gl_account_requires_gen_prod_group() ci-dessous.
#
# Mécanisme BC réel (confirmé par la documentation Microsoft et plusieurs
# cas terrain, pas une liste de libellés à deviner un par un) : un compte
# GL référencé DIRECTEMENT par un champ-compte (sans article pour fournir
# le Gen. Prod. Posting Group) exige que CE COMPTE ait lui-même son
# "Gen. Posting Type" — donc son Groupe compta. produit — rempli, MAIS
# uniquement si ce compte est lui-même typé Vente/Achat (Gen. Posting
# Type non vide). Un compte au Gen. Posting Type vide (ex. Compte client,
# Compte fournisseur — comptes de bilan) n'est jamais concerné. Le Groupe
# compta. marché, lui, vient toujours du client/fournisseur du document,
# jamais du compte GL — structurellement jamais requis ici, ce qui
# explique (sans exception à maintenir) pourquoi il n'a jamais été observé
# en erreur BC pour "Compte frais supplémentaires"/"Compte intérêts".
#
# GL_ACCOUNT_FIELD_REQUIREMENTS est CONSERVÉ comme repli explicite : la
# règle dynamique a besoin du champ "Gen. Posting Type" sur le compte
# (fichier OU live), qui n'est pas garanti disponible partout tout de
# suite (absent du socle MDD Comptabilité actuel, à confirmer côté BC —
# voir GEN_POSTING_TYPE_AL_FIELD). Si ce champ est indisponible pour un
# compte donné, on retombe sur ce mapping connu plutôt que de perdre la
# détection déjà validée le 28/07/2026 sur ces 2 cas réels.
GL_ACCOUNT_FIELD_REQUIREMENTS: dict[str, list[str]] = {
    "compte frais supplémentaires": ["Groupe compta. produit"],
    "compte intérêts": ["Groupe compta. produit"],
}

# Nom de champ AL (interne, stable quelle que soit la langue de la
# société BC) — c'est CE nom qui doit être passé à
# resolve_field_no_via_package() côté bc_api.py pour la résolution live.
# Ne pas confondre avec le libellé français affiché dans le fichier Excel
# (fieldCaption) — voir _GEN_POSTING_TYPE_CAPTION_CANDIDATES ci-dessous.
GEN_POSTING_TYPE_AL_FIELD = "Gen. Posting Type"

# MARQUEUR DE VERSION TEMPORAIRE (04/08/2026) — pour confirmer sans
# ambiguïté ce qui tourne réellement en production, après plusieurs
# décalages de déploiement suspectés sur ce fichier. À retirer une fois
# le problème du "toujours 0 anomalie sur Revérifier" résolu.
CORRECTION_CLASSIFIER_DEBUG_VERSION = "2026-08-04-v4-marker-test"

# NON CONFIRMÉ CONTRE UN VRAI FICHIER BC — le socle MDD Comptabilité
# actuel (vérifié le 04/08/2026 sur Par_défaut28_07_2026_16_42_36.xlsx)
# N'INCLUT PAS ce champ dans son export (51 colonnes présentes, aucune ne
# correspond). Cette liste de libellés candidats sert de détection best-
# effort SI le champ est ajouté un jour à la sélection de champs du
# package — à réduire à la seule vraie valeur dès qu'un export BC réel
# l'inclut et confirme le libellé exact.
_GEN_POSTING_TYPE_CAPTION_CANDIDATES = (
    "type comptabilisation générale",
    "type comptabilisation généraux",
    "gen. posting type",
)


def _gl_account_requires_gen_prod_group(gl_row, gl_columns) -> bool | None:
    """
    True  : le compte a un Gen. Posting Type non vide -> Groupe compta.
            produit doit être rempli sur ce compte.
    False : Gen. Posting Type est présent et vide -> rien à exiger.
    None  : Gen. Posting Type n'est disponible ni dans le fichier ni dans
            le repli (live/cache) -> l'appelant doit se rabattre sur
            GL_ACCOUNT_FIELD_REQUIREMENTS, pas conclure à tort "rien à
            exiger" par absence de donnée.
    """
    gen_col = next(
        (c for c in gl_columns if str(c).strip().lower() in _GEN_POSTING_TYPE_CAPTION_CANDIDATES
         or str(c).strip() == GEN_POSTING_TYPE_AL_FIELD),
        None,
    )
    if gen_col is None:
        return None
    val = gl_row.get(gen_col, "")
    if pd.isna(val):
        return False
    return str(val).strip() not in ("", " ")


def _is_account_reference_column(col_name: str) -> bool:
    """
    Un champ "champ-compte" pointe vers un compte général (Chart of
    Accounts) — reconnu par son libellé BC standard : commence par
    "Compte" ou "Cpte" (Compte client, Compte frais forfaitaires, Cpte
    arrondi débit...), à l'exclusion des champs qui contiennent le mot
    sans être une référence de compte (ex. la case à cocher "Afficher tous
    les comptes lors de la consultation").

    NON VÉRIFIÉ AU-DELÀ DES ONGLETS 92/93 DU FICHIER RÉEL DE RAMI (22/07) :
    la règle générique tient sur ces deux tables, mais n'a pas été testée
    sur d'autres tables de groupes comptables (FA Posting Group, Bank
    Account Posting Group...) si elles apparaissent un jour dans le socle —
    à revérifier si de nouveaux libellés de champ ne matchent pas ce motif.
    """
    name = str(col_name or "").strip().lower()
    if name in _ACCOUNT_FIELD_EXCLUSIONS:
        return False
    return any(name.startswith(p) for p in _ACCOUNT_FIELD_PREFIXES)


def extract_gl_account_posting_fields(parsed_file: dict) -> dict:
    """
    Extrait {"<N° compte>": {"Groupe compta. marché": "...", "Groupe
    compta. produit": "..."}, ...} depuis l'onglet Compte général (table 15)
    d'un fichier déjà parsé — à appeler côté page juste après une analyse
    qui contient réellement cet onglet, pour persister l'état courant via
    app.db.metadata_db.persist_gl_account_posting_fields() (repli utilisé
    par check_gl_account_prerequisites quand un futur fichier n'aura plus
    l'onglet 15, ex. test isolé de 92/93/94).

    Retourne {} si l'onglet Compte général est absent ou vide — rien à
    persister dans ce cas, ne pas écraser une image précédente valide avec
    du vide (l'appelant ne doit persister QUE si ce dict est non vide).
    """
    sheets   = parsed_file.get("sheets", {})
    metadata = parsed_file.get("metadata", {})

    gl_sheet_name = next(
        (name for name, meta in metadata.items() if str(meta.get("table_id", "")) == "15"),
        None,
    )
    if gl_sheet_name is None:
        return {}

    gl_df = sheets.get(gl_sheet_name)
    if gl_df is None or gl_df.empty:
        return {}

    account_col = next((c for c in gl_df.columns if str(c).strip() == "N°"), None)
    if account_col is None:
        return {}

    out = {}
    present_fields = [f for f in GL_ACCOUNT_REQUIRED_FIELDS if f in gl_df.columns]
    # Capture best-effort du Gen. Posting Type si présent dans le fichier
    # (voir _GEN_POSTING_TYPE_CAPTION_CANDIDATES — libellé non confirmé à
    # ce jour, absent du socle MDD Comptabilité actuel). On garde le nom
    # de colonne RÉEL du fichier (pas une constante) pour que
    # _gl_account_requires_gen_prod_group() le retrouve tel quel côté
    # check_gl_account_prerequisites.
    gen_col = next(
        (c for c in gl_df.columns
         if str(c).strip().lower() in _GEN_POSTING_TYPE_CAPTION_CANDIDATES
         or str(c).strip() == GEN_POSTING_TYPE_AL_FIELD),
        None,
    )
    if gen_col is not None:
        present_fields = present_fields + [gen_col]
    if not present_fields:
        return {}

    for _, row in gl_df.iterrows():
        acc_no = str(row.get(account_col, "")).strip()
        if not acc_no:
            continue
        out[acc_no] = {
            f: ("" if pd.isna(row.get(f, "")) else str(row.get(f, "")).strip())
            for f in present_fields
        }
    return out


def check_gl_account_prerequisites(
    parsed_file: dict,
    gl_reference_fallback: dict[str, dict] | None = None,
    prefer_fallback: bool = False,
) -> list[dict]:
    """
    Pour chaque compte général référencé par un champ-compte d'une table de
    groupe comptable (92, 93, 94, ou toute autre table du même type présente
    dans le fichier), vérifie que ce compte a bien ses propres champs
    "Groupe compta. marché" et "Groupe compta. produit" remplis dans
    l'onglet Compte général (table 15) — AVANT de considérer 92/93/94
    intégrables sans erreur BC.

    Ne fait AUCUN appel BC lui-même : contrôle purement inter-onglets sur le
    fichier déjà déposé (parse_uploaded_file) et/ou sur gl_reference_fallback
    fourni par l'appelant (qui, lui, peut avoir interrogé BC en direct — voir
    prefer_fallback ci-dessous). S'utilise en amont de l'intégration BC de
    92/93/94, pas après-coup sur une erreur déjà survenue.

    Retourne une liste de dicts au même format que build_prerequisites_report
    (réutilisable tel quel avec build_prerequisites_excel) :
        {
          "Table référencée BC": "15",
          "Nom table BC": "Compte général",
          "Code manquant": "<N° compte> — <champ vide>",
          "Champs concernés": "<onglet>.<colonne> ; ...",
          "Occurrences": <int>,
        }

    gl_reference_fallback : dict[str, dict] | None
        Repli utilisé quand le fichier déposé ne contient PAS l'onglet
        Compte général (table 15), OU en priorité si prefer_fallback=True
        (voir ci-dessous). Format attendu :
        {"<N° compte>": {"Groupe compta. marché": "...", "Groupe compta.
        produit": "..."}, ...} — typiquement le résultat live de
        app.core.bc_api.get_gl_account_fields_live(), ou à défaut le cache
        Supabase bc_metadata_cache persisté (voir app/db/metadata_db.py,
        entity_name="gl_account_posting_fields").

    prefer_fallback : bool
        AJOUTÉ (28/07/2026) : par défaut False (comportement historique —
        l'onglet du fichier déposé prime toujours). À mettre à True quand
        gl_reference_fallback provient d'une lecture BC live confirmée
        réussie : dans ce cas le live doit primer sur l'onglet du fichier,
        pas l'inverse. Cas confirmé en réel le 28/07/2026 : un fichier
        template repris d'UNE AUTRE société (pour être intégré dans la
        société de test) contient son propre onglet Compte général, sans
        rapport avec l'état réel de la société testée — l'utiliser en
        priorité aurait fait remonter à tort les 2 comptes comme non
        corrigés alors qu'ils l'étaient réellement dans BC, ET aurait
        pollué le cache persisté avec ces valeurs étrangères au prochain
        appel de persist_gl_account_posting_fields côté appelant.
        Si gl_reference_fallback est vide/absent malgré prefer_fallback=True,
        on retombe quand même sur l'onglet du fichier (mieux que rien).

    Si ni l'onglet Compte général du fichier NI gl_reference_fallback ne
    sont disponibles, retourne [] silencieusement (rien à vérifier).
    """
    sheets   = parsed_file.get("sheets", {})
    metadata = parsed_file.get("metadata", {})

    gl_sheet_name = next(
        (name for name, meta in metadata.items() if str(meta.get("table_id", "")) == "15"),
        None,
    )
    gl_df = sheets.get(gl_sheet_name) if gl_sheet_name else None
    if gl_df is not None and gl_df.empty:
        gl_df = None

    gl_by_account = None
    using_fallback = False

    if prefer_fallback and gl_reference_fallback:
        gl_by_account = pd.DataFrame.from_dict(gl_reference_fallback, orient="index")
        using_fallback = True

    if gl_by_account is None and gl_df is not None:
        account_col = next((c for c in gl_df.columns if str(c).strip() == "N°"), None)
        if account_col is not None:
            gl_by_account = gl_df.set_index(gl_df[account_col].astype(str).str.strip())

    # Repli sur l'image fournie (live BC ou cache persisté) si l'onglet du
    # fichier n'est pas exploitable — voir docstring ci-dessus.
    if gl_by_account is None and gl_reference_fallback:
        gl_by_account = pd.DataFrame.from_dict(gl_reference_fallback, orient="index")
        using_fallback = True

    if gl_by_account is None or gl_by_account.empty:
        return []

    missing: dict[tuple, dict] = {}

    for sheet_name, df in sheets.items():
        if sheet_name == gl_sheet_name or df is None or df.empty:
            continue

        account_columns = [c for c in df.columns if _is_account_reference_column(c)]
        if not account_columns:
            continue

        for col in account_columns:
            for raw_value in df[col].dropna():
                acc_no = str(raw_value).strip()
                if not acc_no or acc_no not in gl_by_account.index:
                    continue  # compte inexistant : déjà signalé par ailleurs (Axe B), pas ce contrôle

                gl_row = gl_by_account.loc[acc_no]
                if hasattr(gl_row, "ndim") and gl_row.ndim > 1:
                    gl_row = gl_row.iloc[0]  # N° dupliqué dans le fichier — on ne plante pas, 1re occurrence

                # RÈGLE DYNAMIQUE EN PRIORITÉ (04/08/2026) : déterminée par
                # le Gen. Posting Type du compte lui-même, pas par le nom
                # du champ qui le référence — voir
                # _gl_account_requires_gen_prod_group(). Repli sur le
                # mapping statique connu UNIQUEMENT si cette donnée est
                # indisponible (None), jamais si elle vaut False (compte
                # confirmé non concerné, ne pas écraser cette conclusion
                # par une supposition du mapping statique).
                _dynamic = _gl_account_requires_gen_prod_group(gl_row, gl_by_account.columns)
                # CORRIGÉ (06/08/2026) : la règle dynamique ne doit JAMAIS
                # retirer un cas déjà confirmé par GL_ACCOUNT_FIELD_REQUIREMENTS
                # (mapping validé sur de vraies erreurs BC le 28/07) — seulement
                # en AJOUTER de nouveaux. Bug réel constaté : "Gen. Posting Type"
                # résolu en live avec une valeur vide donnait _dynamic=False
                # (pas None), ce qui effaçait à tort la détection statique
                # confirmée sur 77110001/76310001 (0 anomalie au lieu de 2).
                # Union des deux sources, jamais une substitution.
                required_fields_for_col = set(
                    GL_ACCOUNT_FIELD_REQUIREMENTS.get(str(col).strip().lower(), [])
                )
                if _dynamic is True:
                    required_fields_for_col.add("Groupe compta. produit")
                required_fields_for_col = list(required_fields_for_col)
                for required_field in required_fields_for_col:
                    # ATTENTION : parse_uploaded_file() fait df.dropna(axis=1,
                    # how="all") — une colonne 100% vide sur TOUS les comptes
                    # du fichier disparaît purement et simplement de gl_df.
                    # Donc "colonne absente" ne veut PAS dire "rien à
                    # vérifier" : ça veut dire "vide pour tous les comptes",
                    # exactement le cas confirmé sur le fichier réel de Rami
                    # le 23/07 (aucun compte n'a Groupe compta. marché/
                    # produit rempli). Traiter comme vide, pas comme absent.
                    # Sur le repli (using_fallback), la colonne existe
                    # toujours (construite depuis un dict complet) — même
                    # logique, pas de cas particulier nécessaire.
                    if required_field in gl_by_account.columns:
                        _val = gl_row.get(required_field, "")
                        # BUG CORRIGÉ (24/07) : un NaN pandas (case vide dans
                        # le fichier Excel) est "truthy" en Python — `nan or
                        # ""` vaut nan, pas "". Sans pd.isna() ici, un champ
                        # réellement vide comme celui du compte 40110001
                        # passait à tort pour "rempli". Confirmé en testant
                        # sur le fichier réel de Rami du 24/07 (0 anomalie
                        # remontée alors que Compte fournisseur ETRANGER
                        # avait Groupe compta. marché/produit vides).
                        if not pd.isna(_val) and str(_val).strip():
                            continue  # rempli, rien à signaler

                    key = (acc_no, required_field)
                    if key not in missing:
                        _rule_marker = {
                            True:  "règle dynamique — Gen. Posting Type",
                            False: "règle dynamique — Gen. Posting Type",
                            None:  "mapping statique (Gen. Posting Type indisponible)",
                        }[_dynamic]
                        missing[key] = {
                            "Table référencée BC": "15",
                            "Nom table BC": "Compte général",
                            "Code manquant": f"{acc_no} — {required_field} vide"
                                             + (" (vérifié via socle persisté)" if using_fallback else "")
                                             + f" [{_rule_marker}]",
                            "Champs concernés": set(),
                            "Occurrences": 0,
                        }
                    missing[key]["Champs concernés"].add(f"{sheet_name}.{col}")
                    missing[key]["Occurrences"] += 1

    report = []
    for row in missing.values():
        row["Champs concernés"] = ", ".join(sorted(row["Champs concernés"]))
        report.append(row)
    return sorted(report, key=lambda r: -r["Occurrences"])


def build_prerequisites_excel(prereqs: list[dict]) -> bytes:
    """
    Génère un .xlsx mis en forme (en-tête coloré, colonnes dimensionnées,
    figé sur la 1re ligne) pour la checklist de prérequis BC.

    Remplace le CSV : un CSV en UTF-8 sans BOM s'ouvre en mojibake dans
    Excel FR (accents illisibles, "é" -> "Ã©") tant que l'utilisateur ne
    force pas manuellement l'encodage à l'import. Un .xlsx natif évite le
    problème complètement, et permet la mise en forme demandée.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Prérequis BC"

    ws.append(_PREREQ_COLUMNS)

    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin        = Side(style="thin", color="D1D5DB")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    for row in prereqs:
        ws.append([row.get(col, "") for col in _PREREQ_COLUMNS])

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(_PREREQ_COLUMNS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border    = border
            cell.alignment = Alignment(vertical="center", wrap_text=(c in (3, 4)))
        # Bandes alternées pour la lisibilité
        if r % 2 == 0:
            for c in range(1, len(_PREREQ_COLUMNS) + 1):
                ws.cell(row=r, column=c).fill = PatternFill(
                    start_color="F5F3FF", end_color="F5F3FF", fill_type="solid"
                )

    widths = [18, 26, 34, 34, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()