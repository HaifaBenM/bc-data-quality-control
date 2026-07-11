"""
Traitement du fichier Excel BC Configuration Package.

Préserve intégralement xl/xmlMaps.xml (requis pour import BC)
en ré-injectant les fichiers BC-spécifiques après les modifications openpyxl.
"""
import io
import re
import zipfile

# ── Exemples par table/champ ──────────────────────────────────────────────────
_FIELD_EXAMPLES: dict[int, dict[str, list[str]]] = {
    18: {
        "N°":                      ["CLI-001",    "CLI-002"],
        "Nom":                     ["Société Test SA", "Client Exemple SARL"],
        "Groupe compta. client":   ["NATIONAL",   "ETRANGER"],
        "Groupe compta. marché":   ["NATIONAL",   "NATIONAL"],
        "Code conditions paiement":["30 JOURS",   "60 JOURS"],
        "Code pays/région":        ["FR",         "FR"],
    },
    23: {
        "N°":                      ["FRN-001",    "FRN-002"],
        "Nom":                     ["Fournisseur Test", "Autre Fournisseur"],
        "Groupe compta. fourn.":   ["NATIONAL",   "NATIONAL"],
        "Code conditions paiement":["30 JOURS",   "30 JOURS"],
    },
    27: {
        "N°":                      ["ART-001",    "ART-002"],
        "Description":             ["Article Test 1", "Article Test 2"],
        "Unité de base":           ["PCS",        "PCS"],
        "Groupe compta. stock":    ["MARCHANDISE","MARCHANDISE"],
        "Groupe compta. produit":  ["MARCHANDISE","MARCHANDISE"],
        "Code catégorie article":  ["MATIERES",   "PRODUITS"],
    },
    3:  {"Code": ["30J", "60J"], "Description": ["30 Jours nets", "60 Jours nets"]},
    4:  {"Code": ["EUR", "USD"], "Description": ["Euro",          "Dollar US"]},
    9:  {"Code": ["FR",  "BE"],  "Nom":         ["France",        "Belgique"]},
    15: {"N°":   ["60100", "70100"], "Nom": ["Produits",   "Marchandises"]},
}

_TYPE_EXAMPLES = {
    "Code": "EX-001", "Text": "Exemple", "Decimal": "0.00",
    "Integer": "0", "Date": "01/01/2025", "Boolean": "Non",
}

# ── BC files à préserver ──────────────────────────────────────────────────────
_BC_FILES = {"xl/xmlMaps.xml", "xl/connections.xml"}
_BC_RELS_TYPES = {"xmlMaps", "connections"}


def validate_bc_excel(excel_bytes: bytes) -> tuple[bool, str]:
    try:
        with zipfile.ZipFile(io.BytesIO(excel_bytes)) as z:
            if "xl/xmlMaps.xml" not in z.namelist():
                return False, (
                    "Ce fichier ne contient pas le mapping XML BC (xmlMaps.xml). "
                    "Utilisez le fichier exporté via 'Exporter vers Excel' "
                    "depuis la fiche package dans BC."
                )
        return True, ""
    except zipfile.BadZipFile:
        return False, "Fichier corrompu ou format non reconnu (.xlsx requis)."
    except Exception as e:
        return False, f"Erreur : {e}"


def extract_sheets_info(excel_bytes: bytes) -> list[dict]:
    """Lit les métadonnées et headers de chaque onglet (openpyxl, avant manipulation)."""
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    sheets_info = []
    for ws in wb.worksheets:
        row1  = [ws.cell(1, c).value for c in range(1, 4)]
        table_id   = str(row1[2]).split(".")[0] if row1[2] is not None else ""
        table_name = str(row1[1]) if row1[1] else ws.title
        headers = []
        col = 1
        while True:
            v = ws.cell(3, col).value
            if v is None and col > 5:
                break
            headers.append(str(v) if v is not None else "")
            col += 1
        sheets_info.append({
            "sheet_name": ws.title,
            "table_id":   table_id,
            "table_name": table_name,
            "headers":    headers,
        })
    return sheets_info


def clear_bc_excel_data(excel_bytes: bytes) -> tuple[bytes, list[dict]]:
    """
    Vide les lignes de données (row >= 4) via manipulation zip.
    Retourne (template_bytes, sheets_info).
    """
    sheets_info = extract_sheets_info(excel_bytes)
    input_buf   = io.BytesIO(excel_bytes)
    output_buf  = io.BytesIO()

    with zipfile.ZipFile(input_buf, "r") as zin:
        with zipfile.ZipFile(output_buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if (item.filename.startswith("xl/worksheets/sheet")
                        and item.filename.endswith(".xml")):
                    data = _clear_sheet_rows(data)
                zout.writestr(item, data)

    return output_buf.getvalue(), sheets_info


def _clear_sheet_rows(sheet_xml: bytes) -> bytes:
    """Supprime les lignes r >= 4 du sheetData XML."""
    import xml.etree.ElementTree as ET
    NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    for prefix, uri in re.findall(r'xmlns:?(\w*)="([^"]+)"', sheet_xml.decode("utf-8", errors="replace")):
        try:
            ET.register_namespace(prefix or "", uri)
        except Exception:
            pass
    ET.register_namespace("", NS)
    try:
        root = ET.fromstring(sheet_xml)
        sd = root.find(f"{{{NS}}}sheetData")
        if sd is not None:
            for row in [r for r in sd if int(r.get("r", "0")) >= 4]:
                sd.remove(row)
        return ET.tostring(root, encoding="unicode").encode("utf-8")
    except Exception:
        return sheet_xml


def generate_client_template(
    template_bytes: bytes,
    sheets_info: list[dict],
    options: dict,
) -> bytes:
    """
    Génère le template client avec les options demandées.
    Préserve xmlMaps.xml via ré-injection après modifications openpyxl.

    options:
        include_mandatory : bool  — marque les champs obligatoires (*)
        include_descriptions : bool — ajoute une ligne de description (row 4)
        include_examples : bool  — ajoute 2 lignes d'exemples
        include_custom_fields : bool — ajoute descriptions pour champs perso
    """
    # ── Charger avec openpyxl ─────────────────────────────────────────────────
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))

    for ws, info in zip(wb.worksheets, sheets_info):
        try:
            tid = int(info.get("table_id", 0))
        except (ValueError, TypeError):
            tid = 0

        headers     = info.get("headers", [])
        n_cols      = len(headers)
        next_row    = 4  # ligne courante d'écriture

        try:
            from app.core.validator_axe_a import FIELD_DEFS
            fd_table = FIELD_DEFS.get(str(tid), {})
        except ImportError:
            fd_table = {}

        # ── Option : marquer champs obligatoires ──────────────────────────────
        if options.get("include_mandatory"):
            fill_req  = PatternFill("solid", fgColor="FF1B3A6B")
            fill_std  = PatternFill("solid", fgColor="FF2E6FBF")
            font_w    = Font(bold=True, color="FFFFFFFF", size=10)
            for col_idx, hdr in enumerate(headers, 1):
                fd    = fd_table.get(hdr, {})
                is_req = fd.get("req", False)
                cell  = ws.cell(3, col_idx)
                cell.fill = fill_req if is_req else fill_std
                cell.font = font_w
                if is_req and hdr and not str(hdr).endswith(" *"):
                    cell.value = f"{hdr} *"

        # ── Option : ligne de descriptions ────────────────────────────────────
        if options.get("include_descriptions"):
            # Fond jaune = ligne description (à supprimer avant import BC)
            fill_desc = PatternFill("solid", fgColor="FFFFF3CD")
            font_desc = Font(italic=True, color="FF856404", size=9)
            for col_idx, hdr in enumerate(headers, 1):
                cell = ws.cell(next_row, col_idx)
                cell.fill = fill_desc
                fd = fd_table.get(hdr, {})
                if fd:
                    dtype = fd.get("type", "")
                    maxl  = fd.get("max", "")
                    req   = fd.get("req", False)
                    parts = [f"Type: {dtype}"]
                    if maxl:
                        parts.append(f"Max: {maxl}")
                    if req:
                        parts.append("OBLIGATOIRE")
                    is_custom = col_idx > 200
                    desc = " | ".join(parts)
                    if is_custom and not options.get("include_custom_fields"):
                        desc = ""
                    cell.value = desc
                    cell.font  = font_desc
            next_row += 1

        # ── Option : lignes d'exemples ────────────────────────────────────────
        if options.get("include_examples"):
            tbl_ex  = _FIELD_EXAMPLES.get(tid, {})
            fill_ex = PatternFill("solid", fgColor="FFF0FBF5")
            font_ex = Font(italic=True, color="FF0F6E56", size=9)

            for ex_idx in range(2):  # 2 lignes d'exemple
                # Fond vert = ligne exemple (à supprimer avant import BC)
                for col_idx, hdr in enumerate(headers, 1):
                    cell = ws.cell(next_row, col_idx)
                    cell.fill = fill_ex
                    cell.font = font_ex
                    examples = tbl_ex.get(hdr, [])
                    if ex_idx < len(examples):
                        cell.value = examples[ex_idx]
                    else:
                        fd = fd_table.get(hdr, {})
                        cell.value = _TYPE_EXAMPLES.get(fd.get("type", ""), "")
                next_row += 1

    # ── Sauvegarder avec openpyxl (perd xmlMaps) ─────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    modified_bytes = buf.getvalue()

    # ── Ré-injecter les fichiers BC-spécifiques ───────────────────────────────
    return _reinject_bc_files(modified_bytes, template_bytes)


def _reinject_bc_files(modified_bytes: bytes, original_bytes: bytes) -> bytes:
    """
    Ré-injecte xl/xmlMaps.xml et xl/connections.xml depuis l'original BC,
    et met à jour xl/_rels/workbook.xml.rels + [Content_Types].xml
    pour référencer ces fichiers si openpyxl les a supprimés.
    """
    # Lire fichiers BC depuis l'original
    bc_data: dict[str, bytes] = {}
    orig_rels_str = ""
    orig_ct_str   = ""

    with zipfile.ZipFile(io.BytesIO(original_bytes)) as zorig:
        for fname in _BC_FILES:
            if fname in zorig.namelist():
                bc_data[fname] = zorig.read(fname)
        if "xl/_rels/workbook.xml.rels" in zorig.namelist():
            orig_rels_str = zorig.read("xl/_rels/workbook.xml.rels").decode("utf-8", errors="replace")
        if "[Content_Types].xml" in zorig.namelist():
            orig_ct_str = zorig.read("[Content_Types].xml").decode("utf-8", errors="replace")

    if not bc_data:
        return modified_bytes

    # Extraire les balises Relationship pour xmlMaps et connections
    extra_rels = []
    for rel_tag in re.findall(r'<Relationship[^/]*/>', orig_rels_str):
        if any(t in rel_tag for t in _BC_RELS_TYPES):
            extra_rels.append(rel_tag)

    # Extraire les Override pour xmlMaps et connections
    extra_ct = []
    for ov_tag in re.findall(r'<Override[^/]*/>', orig_ct_str):
        if "xmlMaps" in ov_tag or "connections" in ov_tag.lower():
            extra_ct.append(ov_tag)

    # Rebuilder le zip
    in_buf  = io.BytesIO(modified_bytes)
    out_buf = io.BytesIO()

    with zipfile.ZipFile(in_buf, "r") as zin:
        existing = set(zin.namelist())
        with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)

                if item.filename == "xl/_rels/workbook.xml.rels" and extra_rels:
                    s = data.decode("utf-8", errors="replace")
                    inserts = [r for r in extra_rels if r.split('Type="')[1].split('"')[0].split("/")[-1] not in s]
                    if inserts:
                        s = s.replace("</Relationships>", "\n".join(inserts) + "\n</Relationships>")
                    data = s.encode("utf-8")

                elif item.filename == "[Content_Types].xml" and extra_ct:
                    s = data.decode("utf-8", errors="replace")
                    inserts = [o for o in extra_ct if o.split('PartName="')[1].split('"')[0] not in s]
                    if inserts:
                        s = s.replace("</Types>", "\n".join(inserts) + "\n</Types>")
                    data = s.encode("utf-8")

                zout.writestr(item, data)

            # Ajouter les fichiers BC manquants
            for fname, fdata in bc_data.items():
                if fname not in existing:
                    zout.writestr(fname, fdata)

    return out_buf.getvalue()


def write_corrected_data(
    template_bytes: bytes,
    corrected_dfs:  dict,
) -> bytes:
    from openpyxl.utils import get_column_letter  # lazy import
    """
    Injecte les données corrigées dans le template BC original.

    Préserve intégralement :
      - Onglet Template (instructions Talan)
      - xlmMaps.xml + connections.xml (import BC)
      - Styles, images, commentaires
      - Structure des data sheets (row 1-3)

    Remplace :
      - Row 4+ de chaque data sheet → données corrigées

    Args:
        template_bytes : bytes du template stocké (issu de clear_bc_excel_data)
        corrected_dfs  : {sheet_name: pd.DataFrame} — données corrigées

    Returns:
        bytes du fichier .xlsx BC-compatible prêt à importer
    """
    import xml.etree.ElementTree as ET
    import pandas as pd

    NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    def _df_to_rows_xml(df: pd.DataFrame, start_row: int) -> str:
        """Convertit un DataFrame en XML rows pour sheetData."""
        from openpyxl.utils import get_column_letter
        rows_xml = []
        for r_idx, (_, row) in enumerate(df.iterrows()):
            row_num = start_row + r_idx
            cells = []
            for c_idx, val in enumerate(row, 1):
                col_letter = get_column_letter(c_idx)
                cell_ref   = f"{col_letter}{row_num}"
                if val is None or (isinstance(val, float) and str(val) == 'nan'):
                    val = ""
                val_str = str(val).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
                cells.append(
                    f'<c r="{cell_ref}" t="inlineStr"><is><t>{val_str}</t></is></c>'
                )
            rows_xml.append(f'<row r="{row_num}">{"".join(cells)}</row>')
        return "".join(rows_xml)

    def _inject_data_into_sheet(sheet_xml: bytes, df) -> bytes:
        """
        Vide les rows >= 4 et injecte les données du DataFrame.
        Préserve rows 1-3 (métadonnées + vide + headers).
        """
        # Enregistrer namespaces
        import re
        for pfx, uri in re.findall(r'xmlns:?(\w*)="([^"]+)"', sheet_xml.decode('utf-8', errors='replace')):
            try:
                ET.register_namespace(pfx or "", uri)
            except Exception:
                pass
        ET.register_namespace("", NS)

        try:
            root = ET.fromstring(sheet_xml)
        except ET.ParseError:
            return sheet_xml

        sd = root.find(f"{{{NS}}}sheetData")
        if sd is None:
            return sheet_xml

        # Supprimer rows >= 4
        for row in [r for r in sd if int(r.get("r", "0")) >= 4]:
            sd.remove(row)

        # Injecter les nouvelles données si DataFrame non vide
        if df is not None and not df.empty:
            new_rows_xml = _df_to_rows_xml(df, start_row=4)
            # Parser et ajouter les nouvelles rows
            try:
                wrapper = ET.fromstring(
                    f'<sheetData xmlns="{NS}">{new_rows_xml}</sheetData>'
                )
                for new_row in wrapper:
                    sd.append(new_row)
            except ET.ParseError:
                pass

        return ET.tostring(root, encoding="unicode").encode("utf-8")

    # ── Manipulation zip ─────────────────────────────────────────────────────
    in_buf  = io.BytesIO(template_bytes)
    out_buf = io.BytesIO()

    with zipfile.ZipFile(in_buf, "r") as zin:
        # Map sheet_title → sheet_filename dans le zip
        # Lire workbook.xml pour connaître l'ordre et les noms
        sheet_title_to_file: dict[str, str] = {}
        try:
            wb_xml = zin.read("xl/workbook.xml")
            wb_root = ET.fromstring(wb_xml)
            ns_wb   = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
            sheets_el = wb_root.find(f"{{{ns_wb}}}sheets")
            if sheets_el is not None:
                for i, s in enumerate(sheets_el, 1):
                    title = s.get("name", "")
                    sheet_title_to_file[title] = f"xl/worksheets/sheet{i}.xml"
        except Exception:
            pass

        with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)

                # Chercher si cet onglet a des données corrigées
                matched_df = None
                for title, sheet_file in sheet_title_to_file.items():
                    if item.filename == sheet_file and title in corrected_dfs:
                        matched_df = corrected_dfs[title]
                        break

                if matched_df is not None:
                    # Injecter les données corrigées
                    data = _inject_data_into_sheet(data, matched_df)

                zout.writestr(item, data)

    return out_buf.getvalue()
